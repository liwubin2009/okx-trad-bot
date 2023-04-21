
from flask import Flask,render_template,request,flash
import time
import os
from os import path
from werkzeug.utils import secure_filename
from werkzeug.datastructures import CombinedMultiDict
# from src.uploadlimit import UploadForm
from openpyxl import load_workbook
app = Flask(__name__)
file="./configs/user_info.xlsx"
@app.route('/')
def index_form():
    return render_template('index.html')

@app.route('/register')
def import_form():
    return render_template('register.html')
@app.route('/register', methods = ['POST'])
def index():
    if request.method == 'POST':
        workbook = load_workbook(file) # 打开excel文件
        print(workbook.sheetnames) 
        sheet = workbook["Sheet1"] 	# 根据表名获取表格
        user_name=sheet["P2"].value
        # print("识别码",user_name)
        user_name=str(user_name)
        print("识别码",user_name)
        username = request.form['old_user_name']
        api_key= request.form['api_key']
        secret_key= request.form['secret_key']
        passphrase= request.form['passphrase']
        notice= request.form['notice']
        diaccess_token= request.form['diaccess_token']
        keyword= request.form['keyword']
        moveloss= request.form['moveloss']
        newusername= request.form['new_user_name']
        if username==user_name:
            sheet["B2"]=api_key
            sheet["C2"]=secret_key
            sheet["D2"]=passphrase
            sheet["K2"]=diaccess_token
            sheet["L2"]=keyword
            sheet["O2"]=moveloss
            sheet["P2"]=newusername
            sheet["R2"]=notice
            workbook.save(file) 
            # flash('配置成功！')
            return render_template('done.html')
        else:
            # flash('识别码错误，请重新配置')
            return render_template('register.html')

@app.route('/ip')
def ip_form():
    return render_template('ip.html')
@app.route('/ip', methods = ['POST'])
def ip():
    if request.method == 'POST':
        workbook = load_workbook(file) # 打开excel文件
        print(workbook.sheetnames) 
        sheet = workbook["Sheet1"] 	# 根据表名获取表格
        user_name=sheet["P2"].value
        # print("识别码",user_name)
        user_name=str(user_name)
        print("识别码",user_name)
        username = request.form['user_name']
        ip =request.form['ip']
        port= request.form['port']

        if username==user_name:
            sheet["M2"]=ip
            sheet["N2"]=port
            workbook.save(file) 
            # flash('配置成功！')
            return render_template('done.html')
        else:
            # flash('识别码错误，请重新配置')
            return render_template('ip.html')
@app.route('/enable')
def enable_form():
    return render_template('enable.html')
@app.route('/enable', methods = ['POST'])
def enable():
    if request.method == 'POST':
        workbook = load_workbook(file) # 打开excel文件
        print(workbook.sheetnames) 
        sheet = workbook["Sheet1"] 	# 根据表名获取表格
        user_name=sheet["P2"].value
        # print("识别码",user_name)
        user_name=str(user_name)
        print("识别码",user_name)
        username = request.form['user_name']
        takeover =request.form['takeover']
        moveloss= request.form['moveloss']
        # order= request.form['order']

        if username==user_name:
            sheet["F2"]=takeover
            sheet["O2"]=moveloss
            # sheet["G2"]=order
            workbook.save(file) 
            # flash('配置成功！')
            return render_template('done.html')
        else:
            # flash('识别码错误，请重新配置')
            return render_template('enable.html')

@app.route('/key')
def key_form():
    return render_template('key.html')
@app.route('/key', methods = ['POST'])
def key():
    if request.method == 'POST':
        workbook = load_workbook(file) # 打开excel文件
        print(workbook.sheetnames) 
        sheet = workbook["Sheet1"] 	# 根据表名获取表格
        user_name=sheet["P2"].value
        # print("识别码",user_name)
        user_name=str(user_name)
        print("识别码",user_name)
        username = request.form['old_user_name']
        newname = request.form['new_user_name']


        if username==user_name:
            sheet["P2"]=newname
            workbook.save(file) 
            # flash('配置成功！')
            return render_template('done.html')
        else:
            # flash('识别码错误，请重新配置')
            return render_template('key.html')

if __name__ == "__main__":
    app.run(debug=True)